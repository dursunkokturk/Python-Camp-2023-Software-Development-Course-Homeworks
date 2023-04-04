# Şu ana kadar yaptığımız tüm testleri 
# "Selenium IDE" ile tekrardan ele alınız.

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
from typing import Literal
import openpyxl
from constants import globalConstants

class Test_Demo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()

    def getDataValidLogin():
        excelFile = openpyxl.load_workbook("data/valid_login.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRowsNumber = selectedSheet.max_row
        data = []

        for i in range(2,totalRowsNumber+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data

    def getDataInvalidLogin():
        excelFile = openpyxl.load_workbook("data/invalid_login1_1.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRowsNumber = selectedSheet.max_row
        data = []

        for i in range(2,totalRowsNumber+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
    def getDataLockedOutUser():
        excelFile = openpyxl.load_workbook("data/invalid_login_locked_out_user.xlsx")
        selectedSheet = excelFile["Sayfa1"]

        totalRowsNumber = selectedSheet.max_row
        data = []

        for i in range(2,totalRowsNumber+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
    @pytest.mark.parametrize("username,password",[("","")])
    def test_username_and_password_empty(self, username: Literal[""], password: Literal[""]):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()
        
        self.waitForElementVisible((By.ID, "login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username is required"

        self.driver.save_screenshot(f"{self.folderPath}/test-login-username-and-password-empty-{username}-{password}.png")
    
    @pytest.mark.parametrize("username,password",[("standard_user","")])
    def test_password_empty(self, username: Literal['standard_user'], password: Literal['']):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()
        
        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        
        self.driver.save_screenshot(f"{self.folderPath}/test-login-password-empty-{username}-{password}.png")

        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Password is required"

    @pytest.mark.parametrize("username,password",getDataLockedOutUser())
    def test_locked_out_user(self, username: Literal["locked_out_user"], password: Literal["secret_sauce"]):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()
        
        self.waitForElementVisible((By.ID,"login-button"))
        loginButton = self.driver.find_element(By.ID, "login-button")
        loginButton.click()
        
        self.driver.save_screenshot(f"{self.folderPath}/test-login-locked-out-user-{username}-{password}.png")

        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
    
    @pytest.mark.parametrize("username,password",getDataValidLogin())
    def test_valid_login(self, username: Literal["standard_user"], password: Literal["secret_sauce"]):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()

        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        
        self.driver.save_screenshot(f"{self.folderPath}/test-valid-login-{username}-{password}.png")
    
    @pytest.mark.parametrize("username,password",getDataInvalidLogin())
    def test_invalid_login(self, username: Literal["1"], password: Literal["1"]):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()

        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")

    

    @pytest.mark.parametrize("username,password",getDataInvalidLogin())
    def test_icon_click(self, username: Literal['1'], password: Literal['1']):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()

        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        
        self.driver.save_screenshot(f"{self.folderPath}/test-icon-click-{username}-{password}.png")

        errorIconButton = self.driver.find_element(By.CLASS_NAME,"error-button")
        errorIconButton.click()

    @pytest.mark.parametrize("username,password",getDataValidLogin())
    def test_site_page_redirect(self, username: Literal['standard_user'], password: Literal['secret_sauce']):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()

        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        
        self.driver.get("https://www.saucedemo.com/inventory.html")
        self.driver.save_screenshot(f"{self.folderPath}/test-site-page-redirect-inventory-{username}-{password}.png")
    
    @pytest.mark.parametrize("username,password",getDataValidLogin())
    def test_total_product_number(self, username: Literal["standard_user"], password: Literal["secret_sauce"]):
        self.waitForElementVisible((By.ID,"user-name"))
        loginUserNameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        loginUserPasswordInput = self.driver.find_element(By.ID,"password")
        
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(loginUserNameInput,username)
        actions.send_keys_to_element(loginUserPasswordInput,password)
        actions.perform()

        self.waitForElementVisible((By.ID,"login-button"))
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
        totalProductNumber = self.driver.find_elements(By.CLASS_NAME,"inventory-item")
        
        self.driver.save_screenshot(f"{self.folderPath}/test-total-product-number-{username}-{password}.png")

        print(f"Total Product Number {len(totalProductNumber)}")

    @pytest.mark.parametrize("username,password", getDataValidLogin())
    def test_product_filter(self, username: Literal['standard_user'], password: Literal['secret_sauce']):
        self.waitForElementVisible((By.ID, "user-name"))
        userNameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")

        action = ActionChains(self.driver)
        action.send_keys_to_element(userNameInput, username)
        action.send_keys_to_element(passwordInput, password)
        action.perform()
        logIn = self.driver.find_element(By.ID, "login-button")
        logIn.click()
        filtrele = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[1]/div[2]/div/span/select")
        filtrele.click()
        filtrele2 = self.driver.find_element(By.XPATH,"/html/body/div/div/div/div[1]/div[2]/div/span/select/option[2]")
        self.driver.save_screenshot(f"{self.folderPath}/test-product-filter.png")
        filtrele2.click()

    # sepete urun ekleme ve sepete gitme senaryosu
    @pytest.mark.parametrize("username,password", getDataValidLogin())
    def test_add_product_number_and_go_shopping_cart(self, username: Literal['standard_user'], password: Literal['secret_sauce']):
        self.waitForElementVisible((By.ID, "user-name"))
        loginUserNameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        loginUserPasswordInput = self.driver.find_element(By.ID, "password")

        action = ActionChains(self.driver)
        action.send_keys_to_element(loginUserNameInput, username)
        action.send_keys_to_element(loginUserPasswordInput, password)
        action.perform()
        logIn = self.driver.find_element(By.ID, "login-button")
        logIn.click()

        addToCart = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
        addToCart.click()
        goToCart = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[1]/div[1]/div[3]/a")
        goToCart.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-add-product-number-and-go-shopping-cart.png")

    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(expected_conditions.visibility_of_element_located(locator))